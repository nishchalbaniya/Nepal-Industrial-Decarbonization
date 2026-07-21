"""Excel I/O for plant data."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List
import openpyxl


class ExcelLoader:
    """Load and save plant data from Excel files."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.workbook = openpyxl.load_workbook(path)

    def read_sheet(self, sheet_name: str) -> List[Dict]:
        """Read a sheet as a list of dicts (one per row)."""
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")
        ws = self.workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:]]

    def list_sheets(self) -> List[str]:
        return self.workbook.sheetnames

    def save_results(self, sheet_name: str, data: List[Dict]) -> None:
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        ws = self.workbook.create_sheet(sheet_name)
        if data:
            ws.append(list(data[0].keys()))
            for row in data:
                ws.append(list(row.values()))
        self.workbook.save(self.path)
